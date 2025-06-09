from flask import Flask, request, jsonify, render_template_string, send_file, make_response
from flask_sqlalchemy import SQLAlchemy
import json
import datetime
import os
import io
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

# PostgreSQLè¨­å®š
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL or 'postgresql://localhost:5432/workout_logs'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ãƒ‡ãƒ¼ã‚¿ä¿å­˜ç”¨ãƒ‡ã‚£ãƒ¬ã‚¯ãƒˆãƒªï¼ˆãƒ­ã‚°ç”¨ï¼‰
DATA_DIR = Path("data")
DATA_DIR.mkdir(exist_ok=True)

LOGS_FILE = DATA_DIR / "logs.json"
RECEIVED_DATA_FILE = DATA_DIR / "received_data.json"
CONVERSATIONS_FILE = DATA_DIR / "conversations.json"

# ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ãƒ¢ãƒ‡ãƒ«
class WorkoutSession(db.Model):
    __tablename__ = 'workout_sessions'
    
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    day_of_week = db.Column(db.String(20))
    facility = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)
    
    # ãƒªãƒ¬ãƒ¼ã‚·ãƒ§ãƒ³
    workout_logs = db.relationship('WorkoutLog', backref='session', lazy=True, cascade='all, delete-orphan')

class WorkoutLog(db.Model):
    __tablename__ = 'workout_logs'
    
    id = db.Column(db.Integer, primary_key=True)
    session_id = db.Column(db.Integer, db.ForeignKey('workout_sessions.id'), nullable=False)
    exercise_name = db.Column(db.String(200), nullable=False)
    exercise_category = db.Column(db.String(50))
    weight = db.Column(db.String(50))
    reps = db.Column(db.Integer)
    rest_pause_reps = db.Column(db.Integer, default=0)
    sets = db.Column(db.Integer)
    target_muscle = db.Column(db.String(100))
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.datetime.utcnow)

def log_event(event_type, data=None, status="success", error=None):
    """ã‚¤ãƒ™ãƒ³ãƒˆã‚’ãƒ­ã‚°ã«è¨˜éŒ²"""
    log_entry = {
        "timestamp": datetime.datetime.now().isoformat(),
        "event_type": event_type,
        "status": status,
        "data": data,
        "error": error
    }
    
    # ãƒ­ã‚°ãƒ•ã‚¡ã‚¤ãƒ«ã®èª­ã¿è¾¼ã¿
    logs = []
    if LOGS_FILE.exists():
        try:
            with open(LOGS_FILE, 'r', encoding='utf-8') as f:
                logs = json.load(f)
        except:
            logs = []
    
    logs.append(log_entry)
    
    # æœ€æ–°100ä»¶ã®ã¿ä¿æŒ
    logs = logs[-100:]
    
    # ãƒ­ã‚°ãƒ•ã‚¡ã‚¤ãƒ«ã®æ›¸ãè¾¼ã¿
    with open(LOGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)

def save_received_data(data):
    """å—ä¿¡ã—ãŸãƒ‡ãƒ¼ã‚¿ã‚’ä¿å­˜"""
    received_data = []
    if RECEIVED_DATA_FILE.exists():
        try:
            with open(RECEIVED_DATA_FILE, 'r', encoding='utf-8') as f:
                received_data = json.load(f)
        except:
            received_data = []
    
    data_entry = {
        "timestamp": datetime.datetime.now().isoformat(),
        "data": data
    }
    
    received_data.append(data_entry)
    
    # æœ€æ–°100ä»¶ã®ã¿ä¿æŒ
    received_data = received_data[-100:]
    
    with open(RECEIVED_DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(received_data, f, ensure_ascii=False, indent=2)

def save_conversation_data(data):
    """ä¼šè©±ãƒ‡ãƒ¼ã‚¿ã‚’ä¿å­˜"""
    conversations = []
    if CONVERSATIONS_FILE.exists():
        try:
            with open(CONVERSATIONS_FILE, 'r', encoding='utf-8') as f:
                conversations = json.load(f)
        except:
            conversations = []
    
    conversation_entry = {
        "timestamp": datetime.datetime.now().isoformat(),
        "conversation_id": data.get("conversation_id", f"conv_{len(conversations) + 1}"),
        "data": data
    }
    
    conversations.append(conversation_entry)
    
    # æœ€æ–°100ä»¶ã®ã¿ä¿æŒ
    conversations = conversations[-100:]
    
    with open(CONVERSATIONS_FILE, 'w', encoding='utf-8') as f:
        json.dump(conversations, f, ensure_ascii=False, indent=2)

@app.route('/')
def index():
    """ãƒ¡ã‚¤ãƒ³ãƒšãƒ¼ã‚¸"""
    html = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>GPTs Action Test</title>
        <meta charset="utf-8">
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            .container { max-width: 800px; margin: 0 auto; }
            .section { margin: 20px 0; padding: 20px; border: 1px solid #ddd; border-radius: 5px; }
            .button { background: #007cba; color: white; padding: 10px 20px; text-decoration: none; border-radius: 3px; display: inline-block; margin: 5px; }
            .button:hover { background: #005a8b; }
            .endpoint { background: #f5f5f5; padding: 10px; margin: 10px 0; border-radius: 3px; font-family: monospace; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>GPTs Action Test Application</h1>
            
            <div class="section">
                <h2>ã‚¨ãƒ³ãƒ‰ãƒã‚¤ãƒ³ãƒˆæƒ…å ±</h2>
                <div class="endpoint">
                    <strong>POST:</strong> /api/receive<br>
                    <strong>Content-Type:</strong> application/json<br>
                    <strong>èª¬æ˜:</strong> GPTsã‚¢ã‚¯ã‚·ãƒ§ãƒ³ã‹ã‚‰JSONãƒ‡ãƒ¼ã‚¿ã‚’å—ä¿¡
                </div>
                <div class="endpoint">
                    <strong>POST:</strong> /api/conversation<br>
                    <strong>Content-Type:</strong> application/json<br>
                    <strong>èª¬æ˜:</strong> ä¼šè©±ãƒ‡ãƒ¼ã‚¿ã‚’è§£æãƒ»ä¿å­˜
                </div>
                <div class="endpoint">
                    <strong>POST:</strong> /api/workout<br>
                    <strong>Content-Type:</strong> application/json<br>
                    <strong>èª¬æ˜:</strong> ç­‹ãƒˆãƒ¬ãƒ­ã‚°ã‚’ä¿å­˜ï¼ˆãƒ¬ã‚¹ãƒˆãƒ¬ãƒƒãƒ—æ³•å¯¾å¿œï¼‰
                </div>
            </div>
            
            <div class="section">
                <h2>ç­‹ãƒˆãƒ¬ãƒ­ã‚°</h2>
                <a href="/workouts" class="button">ãƒ¯ãƒ¼ã‚¯ã‚¢ã‚¦ãƒˆä¸€è¦§</a>
                <a href="/workouts/weekly" class="button">é€±æ¬¡ã‚µãƒãƒª</a>
                <a href="/workouts/monthly" class="button">æœˆæ¬¡ã‚µãƒãƒª</a>
                <a href="/api/export/excel" class="button" style="background: #28a745;">ğŸ“Š Excelãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰</a>
            </div>
            
            <div class="section">
                <h2>ãƒ‡ãƒ¼ã‚¿å‚ç…§</h2>
                <a href="/logs" class="button">å‡¦ç†ãƒ­ã‚°ã‚’è¦‹ã‚‹</a>
                <a href="/data" class="button">å—ä¿¡ãƒ‡ãƒ¼ã‚¿ã‚’è¦‹ã‚‹</a>
                <a href="/conversations" class="button">ä¼šè©±ãƒ‡ãƒ¼ã‚¿ã‚’è¦‹ã‚‹</a>
            </div>
            
            <div class="section">
                <h2>ãƒ†ã‚¹ãƒˆç”¨</h2>
                <p>curlã§ãƒ†ã‚¹ãƒˆã™ã‚‹å ´åˆ:</p>
                <div class="endpoint">
curl -X POST {{ request.url_root }}api/receive \\<br>
  -H "Content-Type: application/json" \\<br>
  -d '{"test": "data", "message": "Hello from GPTs"}'
                </div>
            </div>
        </div>
    </body>
    </html>
    '''
    return render_template_string(html)

@app.route('/api/receive', methods=['POST'])
def receive_data():
    """GPTsã‚¢ã‚¯ã‚·ãƒ§ãƒ³ã‹ã‚‰ã®ãƒ‡ãƒ¼ã‚¿ã‚’å—ä¿¡"""
    try:
        # ãƒªã‚¯ã‚¨ã‚¹ãƒˆãƒ‡ãƒ¼ã‚¿ã®å–å¾—
        data = request.get_json()
        
        if data is None:
            log_event("receive_data", error="No JSON data received", status="error")
            return jsonify({"error": "No JSON data received"}), 400
        
        # ãƒ‡ãƒ¼ã‚¿ã®ä¿å­˜
        save_received_data(data)
        
        # ãƒ­ã‚°ã®è¨˜éŒ²
        log_event("receive_data", data=data)
        
        # ãƒ¬ã‚¹ãƒãƒ³ã‚¹
        response_data = {
            "status": "success",
            "message": "Data received and saved",
            "timestamp": datetime.datetime.now().isoformat(),
            "received_data": data
        }
        
        return jsonify(response_data), 200
        
    except Exception as e:
        error_msg = str(e)
        log_event("receive_data", error=error_msg, status="error")
        return jsonify({"error": error_msg}), 500

@app.route('/api/conversation', methods=['POST'])
def save_conversation():
    """ä¼šè©±ãƒ‡ãƒ¼ã‚¿ã‚’å—ä¿¡ãƒ»ä¿å­˜"""
    try:
        # ãƒªã‚¯ã‚¨ã‚¹ãƒˆãƒ‡ãƒ¼ã‚¿ã®å–å¾—
        data = request.get_json()
        
        if data is None:
            log_event("save_conversation", error="No JSON data received", status="error")
            return jsonify({"error": "No JSON data received"}), 400
        
        # å¿…é ˆãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã®ãƒã‚§ãƒƒã‚¯
        if not data.get("user_input") or not data.get("conversation_summary"):
            log_event("save_conversation", error="Missing required fields", status="error")
            return jsonify({"error": "user_input and conversation_summary are required"}), 400
        
        # ä¼šè©±ãƒ‡ãƒ¼ã‚¿ã®ä¿å­˜
        save_conversation_data(data)
        
        # ãƒ­ã‚°ã®è¨˜éŒ²
        log_event("save_conversation", data=data)
        
        # ãƒ¬ã‚¹ãƒãƒ³ã‚¹
        response_data = {
            "status": "success",
            "message": "Conversation data saved successfully",
            "conversation_id": data.get("conversation_id", f"conv_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"),
            "timestamp": datetime.datetime.now().isoformat()
        }
        
        return jsonify(response_data), 200
        
    except Exception as e:
        error_msg = str(e)
        log_event("save_conversation", error=error_msg, status="error")
        return jsonify({"error": error_msg}), 500

@app.route('/api/workout', methods=['POST'])
def save_workout():
    """ç­‹ãƒˆãƒ¬ãƒ­ã‚°ã‚’å—ä¿¡ãƒ»ä¿å­˜"""
    try:
        # ãƒªã‚¯ã‚¨ã‚¹ãƒˆãƒ‡ãƒ¼ã‚¿ã®å–å¾—
        data = request.get_json()
        
        if data is None:
            log_event("save_workout", error="No JSON data received", status="error")
            return jsonify({"error": "No JSON data received"}), 400
        
        # å¿…é ˆãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã®ãƒã‚§ãƒƒã‚¯
        if not data.get("date") or not data.get("exercises"):
            log_event("save_workout", error="Missing required fields", status="error")
            return jsonify({"error": "date and exercises are required"}), 400
        
        # ã‚»ãƒƒã‚·ãƒ§ãƒ³ã‚’ä½œæˆã¾ãŸã¯å–å¾—
        session_date = datetime.datetime.strptime(data['date'], '%Y-%m-%d').date()
        
        session = WorkoutSession.query.filter_by(date=session_date).first()
        if not session:
            session = WorkoutSession(
                date=session_date,
                day_of_week=data.get('day_of_week'),
                facility=data.get('facility')
            )
            db.session.add(session)
            db.session.flush()  # IDã‚’å–å¾—ã™ã‚‹ãŸã‚
        
        # ã‚¨ã‚¯ã‚µã‚µã‚¤ã‚ºãƒ‡ãƒ¼ã‚¿ã‚’ä¿å­˜
        for exercise in data['exercises']:
            workout_log = WorkoutLog(
                session_id=session.id,
                exercise_name=exercise.get('name'),
                exercise_category=exercise.get('category'),
                weight=exercise.get('weight'),
                reps=exercise.get('reps'),
                rest_pause_reps=exercise.get('rest_pause_reps', 0),
                sets=exercise.get('sets'),
                target_muscle=exercise.get('target_muscle'),
                notes=exercise.get('notes')
            )
            db.session.add(workout_log)
        
        db.session.commit()
        
        # ãƒ­ã‚°ã®è¨˜éŒ²
        log_event("save_workout", data={"session_id": session.id, "exercises_count": len(data['exercises'])})
        
        # ãƒ¬ã‚¹ãƒãƒ³ã‚¹
        response_data = {
            "status": "success",
            "message": "Workout data saved successfully",
            "session_id": session.id,
            "date": data['date'],
            "exercises_count": len(data['exercises']),
            "timestamp": datetime.datetime.now().isoformat()
        }
        
        return jsonify(response_data), 200
        
    except Exception as e:
        db.session.rollback()
        error_msg = str(e)
        log_event("save_workout", error=error_msg, status="error")
        return jsonify({"error": error_msg}), 500

@app.route('/api/workout/<int:session_id>', methods=['GET'])
def get_workout_session(session_id):
    """ç‰¹å®šã®ãƒ¯ãƒ¼ã‚¯ã‚¢ã‚¦ãƒˆã‚»ãƒƒã‚·ãƒ§ãƒ³ã‚’å–å¾—"""
    try:
        session = WorkoutSession.query.get_or_404(session_id)
        
        session_data = {
            'id': session.id,
            'date': session.date.strftime('%Y-%m-%d'),
            'day_of_week': session.day_of_week,
            'facility': session.facility,
            'exercises': []
        }
        
        for log in session.workout_logs:
            exercise_data = {
                'id': log.id,
                'name': log.exercise_name,
                'category': log.exercise_category,
                'weight': log.weight,
                'reps': log.reps,
                'rest_pause_reps': log.rest_pause_reps,
                'sets': log.sets,
                'target_muscle': log.target_muscle,
                'notes': log.notes
            }
            session_data['exercises'].append(exercise_data)
        
        return jsonify(session_data), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/workout/<int:session_id>', methods=['DELETE'])
def delete_workout_session(session_id):
    """ãƒ¯ãƒ¼ã‚¯ã‚¢ã‚¦ãƒˆã‚»ãƒƒã‚·ãƒ§ãƒ³ã‚’å‰Šé™¤"""
    try:
        session = WorkoutSession.query.get_or_404(session_id)
        
        # ã‚»ãƒƒã‚·ãƒ§ãƒ³æƒ…å ±ã‚’ãƒ­ã‚°ã«è¨˜éŒ²
        log_event("delete_workout_session", data={
            "session_id": session_id,
            "date": session.date.strftime('%Y-%m-%d'),
            "exercises_count": len(session.workout_logs)
        })
        
        db.session.delete(session)
        db.session.commit()
        
        return jsonify({
            "status": "success",
            "message": "Workout session deleted successfully",
            "session_id": session_id
        }), 200
        
    except Exception as e:
        db.session.rollback()
        error_msg = str(e)
        log_event("delete_workout_session", error=error_msg, status="error")
        return jsonify({"error": error_msg}), 500

@app.route('/api/workout/exercise/<int:exercise_id>', methods=['PUT'])
def update_exercise(exercise_id):
    """å€‹åˆ¥ã‚¨ã‚¯ã‚µã‚µã‚¤ã‚ºã‚’æ›´æ–°"""
    try:
        exercise = WorkoutLog.query.get_or_404(exercise_id)
        data = request.get_json()
        
        if data is None:
            return jsonify({"error": "No JSON data received"}), 400
        
        # ãƒ•ã‚£ãƒ¼ãƒ«ãƒ‰ã‚’æ›´æ–°
        if 'name' in data:
            exercise.exercise_name = data['name']
        if 'category' in data:
            exercise.exercise_category = data['category']
        if 'weight' in data:
            exercise.weight = data['weight']
        if 'reps' in data:
            exercise.reps = data['reps']
        if 'rest_pause_reps' in data:
            exercise.rest_pause_reps = data['rest_pause_reps']
        if 'sets' in data:
            exercise.sets = data['sets']
        if 'target_muscle' in data:
            exercise.target_muscle = data['target_muscle']
        if 'notes' in data:
            exercise.notes = data['notes']
        
        db.session.commit()
        
        log_event("update_exercise", data={"exercise_id": exercise_id, "updated_fields": list(data.keys())})
        
        return jsonify({
            "status": "success",
            "message": "Exercise updated successfully",
            "exercise_id": exercise_id
        }), 200
        
    except Exception as e:
        db.session.rollback()
        error_msg = str(e)
        log_event("update_exercise", error=error_msg, status="error")
        return jsonify({"error": error_msg}), 500

@app.route('/api/workout/exercise/<int:exercise_id>', methods=['DELETE'])
def delete_exercise(exercise_id):
    """å€‹åˆ¥ã‚¨ã‚¯ã‚µã‚µã‚¤ã‚ºã‚’å‰Šé™¤"""
    try:
        exercise = WorkoutLog.query.get_or_404(exercise_id)
        
        log_event("delete_exercise", data={
            "exercise_id": exercise_id,
            "exercise_name": exercise.exercise_name,
            "session_id": exercise.session_id
        })
        
        db.session.delete(exercise)
        db.session.commit()
        
        return jsonify({
            "status": "success",
            "message": "Exercise deleted successfully",
            "exercise_id": exercise_id
        }), 200
        
    except Exception as e:
        db.session.rollback()
        error_msg = str(e)
        log_event("delete_exercise", error=error_msg, status="error")
        return jsonify({"error": error_msg}), 500

def create_excel_export():
    """ç­‹ãƒˆãƒ¬ãƒ­ã‚°ã‚’Excelå½¢å¼ã§å‡ºåŠ›"""
    try:
        # å…¨ã‚»ãƒƒã‚·ãƒ§ãƒ³ã¨ã‚¨ã‚¯ã‚µã‚µã‚¤ã‚ºã‚’å–å¾—ï¼ˆæ—¥ä»˜é †ï¼‰
        sessions = WorkoutSession.query.order_by(WorkoutSession.date.asc()).all()
        
        # Excelãƒ¯ãƒ¼ã‚¯ãƒ–ãƒƒã‚¯ã‚’ä½œæˆ
        wb = Workbook()
        ws = wb.active
        ws.title = "ç­‹ãƒˆãƒ¬ãƒ­ã‚°"
        
        # ãƒ˜ãƒƒãƒ€ãƒ¼è¡Œã®è¨­å®šï¼ˆå…ƒã®Excelãƒ•ã‚¡ã‚¤ãƒ«ã¨åŒã˜æ§‹é€ ï¼‰
        headers = [
            "æ—¥ä»˜", "æ›œæ—¥", "æ–½è¨­å", "ç¨®ç›®", "ç¨®åˆ¥", 
            "é‡é‡(kg)", "å›æ•°(rep)", "ãƒ¬ã‚¹ãƒˆãƒ¬ãƒƒãƒ—", "ã‚»ãƒƒãƒˆæ•°", "éƒ¨ä½", "å‚™è€ƒ"
        ]
        
        # ãƒ˜ãƒƒãƒ€ãƒ¼ã‚¹ã‚¿ã‚¤ãƒ«è¨­å®š
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # ãƒ˜ãƒƒãƒ€ãƒ¼è¡Œã‚’è¨­å®š
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # ãƒ‡ãƒ¼ã‚¿è¡Œã‚’è¿½åŠ 
        row_num = 2
        for session in sessions:
            for exercise in session.workout_logs:
                # ãƒ¬ã‚¹ãƒˆãƒ¬ãƒƒãƒ—å›æ•°ã®è¡¨ç¤ºå½¢å¼
                rest_pause_display = ""
                if exercise.rest_pause_reps and exercise.rest_pause_reps > 0:
                    rest_pause_display = str(exercise.rest_pause_reps)
                
                # ãƒ‡ãƒ¼ã‚¿è¡Œ
                data_row = [
                    session.date.strftime('%Y/%m/%d'),  # æ—¥ä»˜
                    session.day_of_week or "",          # æ›œæ—¥
                    session.facility or "",             # æ–½è¨­å
                    exercise.exercise_name or "",       # ç¨®ç›®
                    exercise.exercise_category or "",   # ç¨®åˆ¥
                    exercise.weight or "",              # é‡é‡(kg)
                    exercise.reps or "",                # å›æ•°(rep)
                    rest_pause_display,                 # ãƒ¬ã‚¹ãƒˆãƒ¬ãƒƒãƒ—
                    exercise.sets or "",                # ã‚»ãƒƒãƒˆæ•°
                    exercise.target_muscle or "",       # éƒ¨ä½
                    exercise.notes or ""                # å‚™è€ƒ
                ]
                
                for col_num, value in enumerate(data_row, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
                
                row_num += 1
        
        # åˆ—å¹…ã®è‡ªå‹•èª¿æ•´
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # æœ€å¤§50æ–‡å­—
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Excelãƒ•ã‚¡ã‚¤ãƒ«ã‚’ãƒ¡ãƒ¢ãƒªã«ä¿å­˜
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer
        
    except Exception as e:
        raise Exception(f"Excel export error: {str(e)}")

@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    """ç­‹ãƒˆãƒ¬ãƒ­ã‚°ã‚’Excelãƒ•ã‚¡ã‚¤ãƒ«ã§ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰"""
    try:
        # Excelãƒ•ã‚¡ã‚¤ãƒ«ã‚’ç”Ÿæˆ
        excel_buffer = create_excel_export()
        
        # ãƒ•ã‚¡ã‚¤ãƒ«åç”Ÿæˆï¼ˆç¾åœ¨ã®æ—¥æ™‚ï¼‰
        filename = f"workout_log_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ç”¨ãƒ¬ã‚¹ãƒãƒ³ã‚¹ã‚’ä½œæˆ
        response = make_response(excel_buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        # ãƒ­ã‚°è¨˜éŒ²
        log_event("export_excel", data={"filename": filename})
        
        return response
        
    except Exception as e:
        error_msg = str(e)
        log_event("export_excel", error=error_msg, status="error")
        return jsonify({"error": error_msg}), 500

@app.route('/workouts')
def view_workouts():
    """ç­‹ãƒˆãƒ¬ãƒ­ã‚°ã®ä¸€è¦§è¡¨ç¤º"""
    try:
        sessions = WorkoutSession.query.order_by(WorkoutSession.date.desc()).limit(30).all()
        
        sessions_data = []
        for session in sessions:
            session_data = {
                'id': session.id,
                'date': session.date.strftime('%Y-%m-%d'),
                'day_of_week': session.day_of_week,
                'facility': session.facility,
                'exercises': []
            }
            
            for log in session.workout_logs:
                exercise_data = {
                    'id': log.id,
                    'name': log.exercise_name,
                    'category': log.exercise_category,
                    'weight': log.weight,
                    'reps': log.reps,
                    'rest_pause_reps': log.rest_pause_reps,
                    'sets': log.sets,
                    'target_muscle': log.target_muscle,
                    'notes': log.notes
                }
                session_data['exercises'].append(exercise_data)
            
            sessions_data.append(session_data)
        
        html = '''
        <!DOCTYPE html>
        <html>
        <head>
            <title>ç­‹ãƒˆãƒ¬ãƒ­ã‚° - GPTs Action Test</title>
            <meta charset="utf-8">
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                .container { max-width: 1200px; margin: 0 auto; }
                .session { margin: 20px 0; padding: 20px; border: 1px solid #ddd; border-radius: 8px; border-left: 4px solid #4CAF50; position: relative; }
                .session-header { background: #f8f9fa; padding: 15px; margin: -20px -20px 15px -20px; border-radius: 8px 8px 0 0; }
                .session-date { font-size: 1.2em; font-weight: bold; color: #2c3e50; }
                .session-info { color: #666; margin-top: 5px; }
                .exercise { margin: 10px 0; padding: 15px; background: #f9f9f9; border-radius: 5px; position: relative; }
                .exercise-header { font-weight: bold; color: #34495e; margin-bottom: 8px; display: flex; justify-content: space-between; align-items: center; }
                .exercise-details { display: grid; grid-template-columns: repeat(auto-fit, minmax(120px, 1fr)); gap: 10px; font-size: 0.9em; }
                .detail-item { background: white; padding: 8px; border-radius: 3px; }
                .detail-label { font-weight: bold; color: #7f8c8d; }
                .detail-value { color: #2c3e50; }
                .notes { margin-top: 10px; padding: 10px; background: #fff3cd; border-radius: 3px; font-style: italic; }
                .button { background: #007cba; color: white; padding: 10px 20px; text-decoration: none; border-radius: 3px; display: inline-block; margin: 5px; border: none; cursor: pointer; }
                .button:hover { background: #005a8b; }
                .button-small { padding: 5px 10px; font-size: 0.8em; margin: 2px; }
                .button-edit { background: #28a745; }
                .button-edit:hover { background: #218838; }
                .button-delete { background: #dc3545; }
                .button-delete:hover { background: #c82333; }
                .session-delete { position: absolute; top: 15px; right: 15px; }
                .exercise-actions { display: flex; gap: 5px; }
                .reps-display { font-weight: bold; }
                .rest-pause { color: #e74c3c; }
                .nav-buttons { margin-bottom: 20px; }
                .stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 20px; }
                .stat-card { background: #f8f9fa; padding: 15px; border-radius: 8px; text-align: center; }
                .stat-value { font-size: 1.5em; font-weight: bold; color: #2c3e50; }
                .stat-label { color: #7f8c8d; margin-top: 5px; }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>ç­‹ãƒˆãƒ¬ãƒ­ã‚°</h1>
                
                <div class="nav-buttons">
                    <a href="/" class="button">â† ãƒ›ãƒ¼ãƒ ã«æˆ»ã‚‹</a>
                    <a href="/workouts/weekly" class="button">é€±æ¬¡ã‚µãƒãƒª</a>
                    <a href="/workouts/monthly" class="button">æœˆæ¬¡ã‚µãƒãƒª</a>
                    <a href="/logs" class="button">ã‚·ã‚¹ãƒ†ãƒ ãƒ­ã‚°</a>
                    <a href="/api/export/excel" class="button" style="background: #28a745;">ğŸ“Š Excelãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰</a>
                </div>
                
                <div class="stats">
                    <div class="stat-card">
                        <div class="stat-value">{{ sessions_data|length }}</div>
                        <div class="stat-label">æœ€è¿‘ã®ã‚»ãƒƒã‚·ãƒ§ãƒ³æ•°</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-value">{{ total_exercises }}</div>
                        <div class="stat-label">ç·ã‚¨ã‚¯ã‚µã‚µã‚¤ã‚ºæ•°</div>
                    </div>
                </div>
                
                {% for session in sessions_data %}
                <div class="session" id="session-{{ session.id }}">
                    <button class="button button-small button-delete session-delete" onclick="deleteSession({{ session.id }})">ã‚»ãƒƒã‚·ãƒ§ãƒ³å‰Šé™¤</button>
                    <div class="session-header">
                        <div class="session-date">{{ session.date }} ({{ session.day_of_week }})</div>
                        <div class="session-info">
                            {% if session.facility %}æ–½è¨­: {{ session.facility }}{% endif %}
                            | ã‚¨ã‚¯ã‚µã‚µã‚¤ã‚ºæ•°: {{ session.exercises|length }}
                        </div>
                    </div>
                    
                    {% for exercise in session.exercises %}
                    <div class="exercise" id="exercise-{{ exercise.id }}">
                        <div class="exercise-header">
                            <span>{{ exercise.name }}</span>
                            <div class="exercise-actions">
                                <button class="button button-small button-edit" onclick="editExercise({{ exercise.id }})">ç·¨é›†</button>
                                <button class="button button-small button-delete" onclick="deleteExercise({{ exercise.id }})">å‰Šé™¤</button>
                            </div>
                        </div>
                        <div class="exercise-details">
                            {% if exercise.category %}
                            <div class="detail-item">
                                <div class="detail-label">ç¨®åˆ¥</div>
                                <div class="detail-value">{{ exercise.category }}</div>
                            </div>
                            {% endif %}
                            {% if exercise.weight %}
                            <div class="detail-item">
                                <div class="detail-label">é‡é‡</div>
                                <div class="detail-value">{{ exercise.weight }}</div>
                            </div>
                            {% endif %}
                            <div class="detail-item">
                                <div class="detail-label">å›æ•°</div>
                                <div class="detail-value reps-display">
                                    {{ exercise.reps }}å›
                                    {% if exercise.rest_pause_reps > 0 %}
                                    <span class="rest-pause">+ {{ exercise.rest_pause_reps }}å›</span>
                                    {% endif %}
                                </div>
                            </div>
                            <div class="detail-item">
                                <div class="detail-label">ã‚»ãƒƒãƒˆæ•°</div>
                                <div class="detail-value">{{ exercise.sets }}ã‚»ãƒƒãƒˆ</div>
                            </div>
                            {% if exercise.target_muscle %}
                            <div class="detail-item">
                                <div class="detail-label">å¯¾è±¡ç­‹è‚‰</div>
                                <div class="detail-value">{{ exercise.target_muscle }}</div>
                            </div>
                            {% endif %}
                        </div>
                        {% if exercise.notes %}
                        <div class="notes">{{ exercise.notes }}</div>
                        {% endif %}
                    </div>
                    {% endfor %}
                </div>
                {% else %}
                <p>ã¾ã ç­‹ãƒˆãƒ¬ãƒ­ã‚°ãŒã‚ã‚Šã¾ã›ã‚“ã€‚</p>
                {% endfor %}
            </div>
            
            <script>
            function deleteSession(sessionId) {
                if (confirm('ã“ã®ã‚»ãƒƒã‚·ãƒ§ãƒ³å…¨ä½“ã‚’å‰Šé™¤ã—ã¾ã™ã‹ï¼Ÿ')) {
                    fetch(`/api/workout/${sessionId}`, {
                        method: 'DELETE'
                    })
                    .then(response => response.json())
                    .then(data => {
                        if (data.status === 'success') {
                            document.getElementById(`session-${sessionId}`).remove();
                            alert('ã‚»ãƒƒã‚·ãƒ§ãƒ³ã‚’å‰Šé™¤ã—ã¾ã—ãŸ');
                            location.reload();
                        } else {
                            alert('å‰Šé™¤ã«å¤±æ•—ã—ã¾ã—ãŸ: ' + data.error);
                        }
                    })
                    .catch(error => {
                        alert('ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: ' + error);
                    });
                }
            }
            
            function deleteExercise(exerciseId) {
                if (confirm('ã“ã®ã‚¨ã‚¯ã‚µã‚µã‚¤ã‚ºã‚’å‰Šé™¤ã—ã¾ã™ã‹ï¼Ÿ')) {
                    fetch(`/api/workout/exercise/${exerciseId}`, {
                        method: 'DELETE'
                    })
                    .then(response => response.json())
                    .then(data => {
                        if (data.status === 'success') {
                            document.getElementById(`exercise-${exerciseId}`).remove();
                            alert('ã‚¨ã‚¯ã‚µã‚µã‚¤ã‚ºã‚’å‰Šé™¤ã—ã¾ã—ãŸ');
                        } else {
                            alert('å‰Šé™¤ã«å¤±æ•—ã—ã¾ã—ãŸ: ' + data.error);
                        }
                    })
                    .catch(error => {
                        alert('ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: ' + error);
                    });
                }
            }
            
            function editExercise(exerciseId) {
                const newReps = prompt('æ–°ã—ã„å›æ•°ã‚’å…¥åŠ›ã—ã¦ãã ã•ã„:');
                if (newReps !== null && newReps !== '') {
                    const updateData = { reps: parseInt(newReps) };
                    
                    fetch(`/api/workout/exercise/${exerciseId}`, {
                        method: 'PUT',
                        headers: {
                            'Content-Type': 'application/json'
                        },
                        body: JSON.stringify(updateData)
                    })
                    .then(response => response.json())
                    .then(data => {
                        if (data.status === 'success') {
                            alert('ã‚¨ã‚¯ã‚µã‚µã‚¤ã‚ºã‚’æ›´æ–°ã—ã¾ã—ãŸ');
                            location.reload();
                        } else {
                            alert('æ›´æ–°ã«å¤±æ•—ã—ã¾ã—ãŸ: ' + data.error);
                        }
                    })
                    .catch(error => {
                        alert('ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: ' + error);
                    });
                }
            }
            </script>
        </body>
        </html>
        '''
        
        total_exercises = sum(len(session['exercises']) for session in sessions_data)
        return render_template_string(html, sessions_data=sessions_data, total_exercises=total_exercises)
        
    except Exception as e:
        return f"ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: {str(e)}", 500

@app.route('/workouts/weekly')
def view_weekly_summary():
    """é€±æ¬¡ã‚µãƒãƒªè¡¨ç¤º"""
    try:
        from sqlalchemy import func, text
        
        # éå»8é€±é–“ã®ãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—
        weekly_data = db.session.query(
            func.date_trunc('week', WorkoutSession.date).label('week_start'),
            func.count(WorkoutSession.id).label('session_count'),
            func.count(WorkoutLog.id).label('exercise_count')
        ).join(WorkoutLog).group_by(
            func.date_trunc('week', WorkoutSession.date)
        ).order_by(
            func.date_trunc('week', WorkoutSession.date).desc()
        ).limit(8).all()
        
        # ç­‹è‚‰éƒ¨ä½åˆ¥ã®é€±æ¬¡çµ±è¨ˆ
        muscle_stats = db.session.query(
            WorkoutLog.target_muscle,
            func.count(WorkoutLog.id).label('exercise_count'),
            func.date_trunc('week', WorkoutSession.date).label('week_start')
        ).join(WorkoutSession).filter(
            WorkoutLog.target_muscle.isnot(None)
        ).group_by(
            WorkoutLog.target_muscle,
            func.date_trunc('week', WorkoutSession.date)
        ).order_by(
            func.date_trunc('week', WorkoutSession.date).desc()
        ).limit(50).all()
        
        html = '''
        <!DOCTYPE html>
        <html>
        <head>
            <title>é€±æ¬¡ã‚µãƒãƒª - ç­‹ãƒˆãƒ¬ãƒ­ã‚°</title>
            <meta charset="utf-8">
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                .container { max-width: 1200px; margin: 0 auto; }
                .week-summary { margin: 20px 0; padding: 20px; border: 1px solid #ddd; border-radius: 8px; border-left: 4px solid #2196F3; }
                .week-header { font-size: 1.2em; font-weight: bold; color: #2c3e50; margin-bottom: 15px; }
                .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 15px; }
                .stat-card { background: #f8f9fa; padding: 15px; border-radius: 8px; text-align: center; }
                .stat-value { font-size: 1.5em; font-weight: bold; color: #2c3e50; }
                .stat-label { color: #7f8c8d; margin-top: 5px; font-size: 0.9em; }
                .muscle-stats { margin-top: 15px; }
                .muscle-item { display: inline-block; background: #e3f2fd; padding: 5px 10px; margin: 3px; border-radius: 15px; font-size: 0.9em; }
                .button { background: #007cba; color: white; padding: 10px 20px; text-decoration: none; border-radius: 3px; display: inline-block; margin: 5px; }
                .button:hover { background: #005a8b; }
                .nav-buttons { margin-bottom: 20px; }
                .chart-container { background: white; padding: 20px; border-radius: 8px; margin: 20px 0; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>é€±æ¬¡ã‚µãƒãƒª</h1>
                
                <div class="nav-buttons">
                    <a href="/workouts" class="button">â† ãƒ¯ãƒ¼ã‚¯ã‚¢ã‚¦ãƒˆä¸€è¦§</a>
                    <a href="/workouts/monthly" class="button">æœˆæ¬¡ã‚µãƒãƒª</a>
                    <a href="/" class="button">ãƒ›ãƒ¼ãƒ </a>
                </div>
                
                {% for week in weekly_data %}
                <div class="week-summary">
                    <div class="week-header">{{ week.week_start.strftime('%Yå¹´%mæœˆ%dæ—¥') }}ã®é€±</div>
                    <div class="stats-grid">
                        <div class="stat-card">
                            <div class="stat-value">{{ week.session_count }}</div>
                            <div class="stat-label">ãƒˆãƒ¬ãƒ¼ãƒ‹ãƒ³ã‚°æ—¥æ•°</div>
                        </div>
                        <div class="stat-card">
                            <div class="stat-value">{{ week.exercise_count }}</div>
                            <div class="stat-label">ç·ã‚¨ã‚¯ã‚µã‚µã‚¤ã‚ºæ•°</div>
                        </div>
                        <div class="stat-card">
                            <div class="stat-value">{{ "%.1f"|format(week.exercise_count / week.session_count if week.session_count > 0 else 0) }}</div>
                            <div class="stat-label">1æ—¥å¹³å‡ã‚¨ã‚¯ã‚µã‚µã‚¤ã‚ºæ•°</div>
                        </div>
                    </div>
                    
                    <div class="muscle-stats">
                        <strong>å¯¾è±¡ç­‹è‚‰:</strong>
                        {% for muscle in muscle_stats %}
                            {% if muscle.week_start == week.week_start %}
                            <span class="muscle-item">{{ muscle.target_muscle }} ({{ muscle.exercise_count }}å›)</span>
                            {% endif %}
                        {% endfor %}
                    </div>
                </div>
                {% else %}
                <p>ã¾ã é€±æ¬¡ãƒ‡ãƒ¼ã‚¿ãŒã‚ã‚Šã¾ã›ã‚“ã€‚</p>
                {% endfor %}
            </div>
        </body>
        </html>
        '''
        
        return render_template_string(html, weekly_data=weekly_data, muscle_stats=muscle_stats)
        
    except Exception as e:
        return f"ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: {str(e)}", 500

@app.route('/workouts/monthly')
def view_monthly_summary():
    """æœˆæ¬¡ã‚µãƒãƒªè¡¨ç¤º"""
    try:
        from sqlalchemy import func, extract
        
        # éå»6ãƒ¶æœˆã®ãƒ‡ãƒ¼ã‚¿ã‚’å–å¾—
        monthly_data = db.session.query(
            extract('year', WorkoutSession.date).label('year'),
            extract('month', WorkoutSession.date).label('month'),
            func.count(WorkoutSession.id).label('session_count'),
            func.count(WorkoutLog.id).label('exercise_count'),
            func.count(func.distinct(WorkoutLog.exercise_name)).label('unique_exercises')
        ).join(WorkoutLog).group_by(
            extract('year', WorkoutSession.date),
            extract('month', WorkoutSession.date)
        ).order_by(
            extract('year', WorkoutSession.date).desc(),
            extract('month', WorkoutSession.date).desc()
        ).limit(6).all()
        
        # æœˆåˆ¥ã®ç­‹è‚‰éƒ¨ä½çµ±è¨ˆ
        monthly_muscle_stats = db.session.query(
            extract('year', WorkoutSession.date).label('year'),
            extract('month', WorkoutSession.date).label('month'),
            WorkoutLog.target_muscle,
            func.count(WorkoutLog.id).label('exercise_count')
        ).join(WorkoutSession).filter(
            WorkoutLog.target_muscle.isnot(None)
        ).group_by(
            extract('year', WorkoutSession.date),
            extract('month', WorkoutSession.date),
            WorkoutLog.target_muscle
        ).order_by(
            extract('year', WorkoutSession.date).desc(),
            extract('month', WorkoutSession.date).desc(),
            func.count(WorkoutLog.id).desc()
        ).all()
        
        html = '''
        <!DOCTYPE html>
        <html>
        <head>
            <title>æœˆæ¬¡ã‚µãƒãƒª - ç­‹ãƒˆãƒ¬ãƒ­ã‚°</title>
            <meta charset="utf-8">
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                .container { max-width: 1200px; margin: 0 auto; }
                .month-summary { margin: 20px 0; padding: 20px; border: 1px solid #ddd; border-radius: 8px; border-left: 4px solid #FF9800; }
                .month-header { font-size: 1.3em; font-weight: bold; color: #2c3e50; margin-bottom: 15px; }
                .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 15px; margin-bottom: 15px; }
                .stat-card { background: #f8f9fa; padding: 15px; border-radius: 8px; text-align: center; }
                .stat-value { font-size: 1.5em; font-weight: bold; color: #2c3e50; }
                .stat-label { color: #7f8c8d; margin-top: 5px; font-size: 0.9em; }
                .muscle-stats { margin-top: 15px; }
                .muscle-item { display: inline-block; background: #fff3e0; padding: 5px 10px; margin: 3px; border-radius: 15px; font-size: 0.9em; }
                .button { background: #007cba; color: white; padding: 10px 20px; text-decoration: none; border-radius: 3px; display: inline-block; margin: 5px; }
                .button:hover { background: #005a8b; }
                .nav-buttons { margin-bottom: 20px; }
                .progress-bar { background: #e0e0e0; height: 10px; border-radius: 5px; margin: 10px 0; }
                .progress-fill { background: #4CAF50; height: 100%; border-radius: 5px; transition: width 0.3s ease; }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>æœˆæ¬¡ã‚µãƒãƒª</h1>
                
                <div class="nav-buttons">
                    <a href="/workouts" class="button">â† ãƒ¯ãƒ¼ã‚¯ã‚¢ã‚¦ãƒˆä¸€è¦§</a>
                    <a href="/workouts/weekly" class="button">é€±æ¬¡ã‚µãƒãƒª</a>
                    <a href="/" class="button">ãƒ›ãƒ¼ãƒ </a>
                </div>
                
                {% for month in monthly_data %}
                <div class="month-summary">
                    <div class="month-header">{{ month.year }}å¹´{{ month.month }}æœˆ</div>
                    <div class="stats-grid">
                        <div class="stat-card">
                            <div class="stat-value">{{ month.session_count }}</div>
                            <div class="stat-label">ãƒˆãƒ¬ãƒ¼ãƒ‹ãƒ³ã‚°æ—¥æ•°</div>
                        </div>
                        <div class="stat-card">
                            <div class="stat-value">{{ month.exercise_count }}</div>
                            <div class="stat-label">ç·ã‚¨ã‚¯ã‚µã‚µã‚¤ã‚ºæ•°</div>
                        </div>
                        <div class="stat-card">
                            <div class="stat-value">{{ month.unique_exercises }}</div>
                            <div class="stat-label">ãƒ¦ãƒ‹ãƒ¼ã‚¯ç¨®ç›®æ•°</div>
                        </div>
                        <div class="stat-card">
                            <div class="stat-value">{{ "%.1f"|format(month.exercise_count / month.session_count if month.session_count > 0 else 0) }}</div>
                            <div class="stat-label">1æ—¥å¹³å‡ã‚¨ã‚¯ã‚µã‚µã‚¤ã‚ºæ•°</div>
                        </div>
                    </div>
                    
                    <div class="muscle-stats">
                        <strong>ä¸»è¦å¯¾è±¡ç­‹è‚‰ TOP5:</strong>
                        {% set month_muscles = [] %}
                        {% for muscle in monthly_muscle_stats %}
                            {% if muscle.year == month.year and muscle.month == month.month %}
                                {% set _ = month_muscles.append(muscle) %}
                            {% endif %}
                        {% endfor %}
                        {% for muscle in month_muscles[:5] %}
                        <span class="muscle-item">{{ muscle.target_muscle }} ({{ muscle.exercise_count }}å›)</span>
                        {% endfor %}
                    </div>
                </div>
                {% else %}
                <p>ã¾ã æœˆæ¬¡ãƒ‡ãƒ¼ã‚¿ãŒã‚ã‚Šã¾ã›ã‚“ã€‚</p>
                {% endfor %}
            </div>
        </body>
        </html>
        '''
        
        return render_template_string(html, monthly_data=monthly_data, monthly_muscle_stats=monthly_muscle_stats)
        
    except Exception as e:
        return f"ã‚¨ãƒ©ãƒ¼ãŒç™ºç”Ÿã—ã¾ã—ãŸ: {str(e)}", 500

@app.route('/logs')
def view_logs():
    """å‡¦ç†ãƒ­ã‚°ã®è¡¨ç¤º"""
    logs = []
    if LOGS_FILE.exists():
        try:
            with open(LOGS_FILE, 'r', encoding='utf-8') as f:
                logs = json.load(f)
        except:
            logs = []
    
    # æœ€æ–°ã®ãƒ­ã‚°ã‹ã‚‰è¡¨ç¤º
    logs.reverse()
    
    html = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>å‡¦ç†ãƒ­ã‚° - GPTs Action Test</title>
        <meta charset="utf-8">
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            .container { max-width: 1000px; margin: 0 auto; }
            .log-entry { margin: 10px 0; padding: 15px; border: 1px solid #ddd; border-radius: 5px; }
            .log-success { border-left: 4px solid #4CAF50; }
            .log-error { border-left: 4px solid #f44336; }
            .timestamp { color: #666; font-size: 0.9em; }
            .data { background: #f5f5f5; padding: 10px; margin: 10px 0; border-radius: 3px; font-family: monospace; }
            .button { background: #007cba; color: white; padding: 10px 20px; text-decoration: none; border-radius: 3px; display: inline-block; margin: 5px; }
            .button:hover { background: #005a8b; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>å‡¦ç†ãƒ­ã‚°</h1>
            <p><a href="/" class="button">â† ãƒ›ãƒ¼ãƒ ã«æˆ»ã‚‹</a> <a href="/data" class="button">å—ä¿¡ãƒ‡ãƒ¼ã‚¿ã‚’è¦‹ã‚‹</a> <a href="/conversations" class="button">ä¼šè©±ãƒ‡ãƒ¼ã‚¿ã‚’è¦‹ã‚‹</a></p>
            
            {% if logs %}
                {% for log in logs %}
                <div class="log-entry {{ 'log-success' if log.status == 'success' else 'log-error' }}">
                    <div class="timestamp">{{ log.timestamp }}</div>
                    <div><strong>ã‚¤ãƒ™ãƒ³ãƒˆ:</strong> {{ log.event_type }}</div>
                    <div><strong>ã‚¹ãƒ†ãƒ¼ã‚¿ã‚¹:</strong> {{ log.status }}</div>
                    {% if log.error %}
                    <div><strong>ã‚¨ãƒ©ãƒ¼:</strong> {{ log.error }}</div>
                    {% endif %}
                    {% if log.data %}
                    <div><strong>ãƒ‡ãƒ¼ã‚¿:</strong></div>
                    <div class="data">{{ log.data | tojson(indent=2) }}</div>
                    {% endif %}
                </div>
                {% endfor %}
            {% else %}
                <p>ã¾ã ãƒ­ã‚°ãŒã‚ã‚Šã¾ã›ã‚“ã€‚</p>
            {% endif %}
        </div>
    </body>
    </html>
    '''
    return render_template_string(html, logs=logs)

@app.route('/data')
def view_data():
    """å—ä¿¡ãƒ‡ãƒ¼ã‚¿ã®è¡¨ç¤º"""
    received_data = []
    if RECEIVED_DATA_FILE.exists():
        try:
            with open(RECEIVED_DATA_FILE, 'r', encoding='utf-8') as f:
                received_data = json.load(f)
        except:
            received_data = []
    
    # æœ€æ–°ã®ãƒ‡ãƒ¼ã‚¿ã‹ã‚‰è¡¨ç¤º
    received_data.reverse()
    
    html = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>å—ä¿¡ãƒ‡ãƒ¼ã‚¿ - GPTs Action Test</title>
        <meta charset="utf-8">
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            .container { max-width: 1000px; margin: 0 auto; }
            .data-entry { margin: 10px 0; padding: 15px; border: 1px solid #ddd; border-radius: 5px; border-left: 4px solid #2196F3; }
            .timestamp { color: #666; font-size: 0.9em; }
            .data { background: #f5f5f5; padding: 10px; margin: 10px 0; border-radius: 3px; font-family: monospace; }
            .button { background: #007cba; color: white; padding: 10px 20px; text-decoration: none; border-radius: 3px; display: inline-block; margin: 5px; }
            .button:hover { background: #005a8b; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>å—ä¿¡ãƒ‡ãƒ¼ã‚¿</h1>
            <p><a href="/" class="button">â† ãƒ›ãƒ¼ãƒ ã«æˆ»ã‚‹</a> <a href="/logs" class="button">å‡¦ç†ãƒ­ã‚°ã‚’è¦‹ã‚‹</a> <a href="/conversations" class="button">ä¼šè©±ãƒ‡ãƒ¼ã‚¿ã‚’è¦‹ã‚‹</a></p>
            
            {% if received_data %}
                {% for entry in received_data %}
                <div class="data-entry">
                    <div class="timestamp">å—ä¿¡æ™‚åˆ»: {{ entry.timestamp }}</div>
                    <div class="data">{{ entry.data | tojson(indent=2) }}</div>
                </div>
                {% endfor %}
            {% else %}
                <p>ã¾ã å—ä¿¡ãƒ‡ãƒ¼ã‚¿ãŒã‚ã‚Šã¾ã›ã‚“ã€‚</p>
            {% endif %}
        </div>
    </body>
    </html>
    '''
    return render_template_string(html, received_data=received_data)

@app.route('/conversations')
def view_conversations():
    """ä¼šè©±ãƒ‡ãƒ¼ã‚¿ã®è¡¨ç¤º"""
    conversations = []
    if CONVERSATIONS_FILE.exists():
        try:
            with open(CONVERSATIONS_FILE, 'r', encoding='utf-8') as f:
                conversations = json.load(f)
        except:
            conversations = []
    
    # æœ€æ–°ã®ä¼šè©±ã‹ã‚‰è¡¨ç¤º
    conversations.reverse()
    
    html = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>ä¼šè©±ãƒ‡ãƒ¼ã‚¿ - GPTs Action Test</title>
        <meta charset="utf-8">
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            .container { max-width: 1200px; margin: 0 auto; }
            .conversation-entry { margin: 15px 0; padding: 20px; border: 1px solid #ddd; border-radius: 8px; border-left: 4px solid #9C27B0; }
            .timestamp { color: #666; font-size: 0.9em; margin-bottom: 10px; }
            .conversation-id { color: #9C27B0; font-weight: bold; margin-bottom: 10px; }
            .section { margin: 10px 0; }
            .section-title { font-weight: bold; color: #333; margin-bottom: 5px; }
            .user-input { background: #e3f2fd; padding: 10px; border-radius: 5px; margin: 5px 0; }
            .assistant-response { background: #f3e5f5; padding: 10px; border-radius: 5px; margin: 5px 0; }
            .summary { background: #fff3e0; padding: 10px; border-radius: 5px; margin: 5px 0; }
            .topics { background: #e8f5e8; padding: 10px; border-radius: 5px; margin: 5px 0; }
            .sentiment { display: inline-block; padding: 3px 8px; border-radius: 12px; font-size: 0.8em; }
            .sentiment-positive { background: #4CAF50; color: white; }
            .sentiment-negative { background: #f44336; color: white; }
            .sentiment-neutral { background: #757575; color: white; }
            .metadata { background: #f5f5f5; padding: 10px; border-radius: 5px; margin: 5px 0; font-family: monospace; font-size: 0.9em; }
            .button { background: #007cba; color: white; padding: 10px 20px; text-decoration: none; border-radius: 3px; display: inline-block; margin: 5px; }
            .button:hover { background: #005a8b; }
            .tags { margin: 5px 0; }
            .tag { background: #e0e0e0; padding: 2px 6px; border-radius: 3px; font-size: 0.8em; margin-right: 5px; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>ä¼šè©±ãƒ‡ãƒ¼ã‚¿</h1>
            <p><a href="/" class="button">â† ãƒ›ãƒ¼ãƒ ã«æˆ»ã‚‹</a> <a href="/logs" class="button">å‡¦ç†ãƒ­ã‚°ã‚’è¦‹ã‚‹</a> <a href="/data" class="button">å—ä¿¡ãƒ‡ãƒ¼ã‚¿ã‚’è¦‹ã‚‹</a></p>
            
            {% if conversations %}
                {% for conversation in conversations %}
                <div class="conversation-entry">
                    <div class="timestamp">è¨˜éŒ²æ™‚åˆ»: {{ conversation.timestamp }}</div>
                    <div class="conversation-id">ä¼šè©±ID: {{ conversation.conversation_id }}</div>
                    
                    {% if conversation.data.user_input %}
                    <div class="section">
                        <div class="section-title">ãƒ¦ãƒ¼ã‚¶ãƒ¼å…¥åŠ›:</div>
                        <div class="user-input">{{ conversation.data.user_input }}</div>
                    </div>
                    {% endif %}
                    
                    {% if conversation.data.assistant_response %}
                    <div class="section">
                        <div class="section-title">ã‚¢ã‚·ã‚¹ã‚¿ãƒ³ãƒˆå›ç­”:</div>
                        <div class="assistant-response">{{ conversation.data.assistant_response }}</div>
                    </div>
                    {% endif %}
                    
                    {% if conversation.data.conversation_summary %}
                    <div class="section">
                        <div class="section-title">ä¼šè©±è¦ç´„:</div>
                        <div class="summary">{{ conversation.data.conversation_summary }}</div>
                    </div>
                    {% endif %}
                    
                    {% if conversation.data.key_topics %}
                    <div class="section">
                        <div class="section-title">ä¸»è¦ãƒˆãƒ”ãƒƒã‚¯:</div>
                        <div class="topics tags">
                            {% for topic in conversation.data.key_topics %}
                            <span class="tag">{{ topic }}</span>
                            {% endfor %}
                        </div>
                    </div>
                    {% endif %}
                    
                    {% if conversation.data.sentiment %}
                    <div class="section">
                        <div class="section-title">æ„Ÿæƒ…å‚¾å‘:</div>
                        <span class="sentiment sentiment-{{ conversation.data.sentiment }}">{{ conversation.data.sentiment }}</span>
                    </div>
                    {% endif %}
                    
                    {% if conversation.data.category %}
                    <div class="section">
                        <div class="section-title">ã‚«ãƒ†ã‚´ãƒª:</div>
                        <span class="tag">{{ conversation.data.category }}</span>
                    </div>
                    {% endif %}
                    
                    {% if conversation.data.action_items %}
                    <div class="section">
                        <div class="section-title">ã‚¢ã‚¯ã‚·ãƒ§ãƒ³ã‚¢ã‚¤ãƒ†ãƒ :</div>
                        <ul>
                            {% for item in conversation.data.action_items %}
                            <li>{{ item }}</li>
                            {% endfor %}
                        </ul>
                    </div>
                    {% endif %}
                    
                    {% if conversation.data.entities %}
                    <div class="section">
                        <div class="section-title">æŠ½å‡ºã‚¨ãƒ³ãƒ†ã‚£ãƒ†ã‚£:</div>
                        <div class="tags">
                            {% for entity in conversation.data.entities %}
                            <span class="tag">{{ entity.type }}: {{ entity.value }}</span>
                            {% endfor %}
                        </div>
                    </div>
                    {% endif %}
                    
                    {% if conversation.data.metadata %}
                    <div class="section">
                        <div class="section-title">ãƒ¡ã‚¿ãƒ‡ãƒ¼ã‚¿:</div>
                        <div class="metadata">{{ conversation.data.metadata | tojson(indent=2) }}</div>
                    </div>
                    {% endif %}
                </div>
                {% endfor %}
            {% else %}
                <p>ã¾ã ä¼šè©±ãƒ‡ãƒ¼ã‚¿ãŒã‚ã‚Šã¾ã›ã‚“ã€‚</p>
            {% endif %}
        </div>
    </body>
    </html>
    '''
    return render_template_string(html, conversations=conversations)

def create_tables():
    """ãƒ‡ãƒ¼ã‚¿ãƒ™ãƒ¼ã‚¹ãƒ†ãƒ¼ãƒ–ãƒ«ã‚’ä½œæˆ"""
    with app.app_context():
        db.create_all()

# ã‚¢ãƒ—ãƒªã‚±ãƒ¼ã‚·ãƒ§ãƒ³èµ·å‹•æ™‚ã«ãƒ†ãƒ¼ãƒ–ãƒ«ä½œæˆ
create_tables()

if __name__ == '__main__':
    log_event("app_start", data={"message": "Application started"})
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)